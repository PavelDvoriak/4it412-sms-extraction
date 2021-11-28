import openpyxl as op

def output_to_xls(data: dict, xls_path=None) -> None:
    wb = op.Workbook()

    ws = wb.active

    for index, key in enumerate_by2(data, 1):
        ws.cell(1,index).value = key
        ws.merge_cells(start_row=1, end_row=1, start_column=index, end_column=index+1)
        for idx, inner_key in enumerate(data[key], 2):
            ws.cell(idx, index).value = inner_key
            ws.cell(idx, index+1).value = data[key][inner_key]

           

    if xls_path:
        wb.save(f'{xls_path}.xlsx')
    else:    
        wb.save('result.xlsx')

def enumerate_by2(sequence, start=0):
     n = start
     for elem in sequence:
         yield n, elem
         n += 2